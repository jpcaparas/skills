#!/usr/bin/env python3
"""Perform a lightweight visual-risk scan on a generated workbook."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Workbook path to inspect.")
    parser.add_argument("--max-findings", type=int, default=25, help="Maximum visual-risk findings to emit.")
    return parser.parse_args()


def cell_text(value) -> str:
    return "" if value is None else str(value)


def scan_visual_risks(workbook_path: Path, max_findings: int) -> dict[str, object]:
    workbook_path = workbook_path.expanduser()
    wb = load_workbook(workbook_path, data_only=False)
    findings: list[dict[str, object]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text = cell_text(cell.value).strip()
                if not text:
                    continue
                column_letter = cell.column_letter
                width = ws.column_dimensions[column_letter].width or 10
                wrap = bool(cell.alignment.wrap_text)
                if len(text) > width * (5.5 if wrap else 1.8):
                    findings.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "issue": "possible_clipping",
                            "text_length": len(text),
                            "column_width": width,
                        }
                    )
                if len(findings) >= max_findings:
                    break
            if len(findings) >= max_findings:
                break
        if len(findings) >= max_findings:
            break

    return {
        "workbook": str(workbook_path),
        "sheet_count": len(wb.sheetnames),
        "findings": findings,
        "render_guidance": "If @oai/artifact-tool or another spreadsheet renderer is available, render each sheet for a final visual pass after this structural scan.",
    }


def main() -> None:
    args = parse_args()
    result = scan_visual_risks(Path(args.workbook), args.max_findings)
    print(json.dumps(result, indent=2))
    raise SystemExit(0)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # pragma: no cover - CLI guard
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
