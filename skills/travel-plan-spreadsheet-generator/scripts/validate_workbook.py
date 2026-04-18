#!/usr/bin/env python3
"""Validate a generated travel workbook against the deterministic contract."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from build_workbook import (
    BLOCKS,
    HEADERS,
    MERGE_SPECS,
    WIDTH_SPECS,
    destination_display_name,
    normalize_trip_model,
    runtime_sheet_specs,
    workbook_sheet_order,
)

WIDTH_TOLERANCE = 0.15


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Workbook path to validate.")
    parser.add_argument("--trip-model", help="Optional trip model JSON path for dynamic-name validation.")
    parser.add_argument("--json", action="store_true", help="Emit JSON only.")
    return parser.parse_args()


def style_error(message: str) -> str:
    return message


def validate_workbook(workbook_path: Path, trip_model_path: Path | None = None) -> dict[str, object]:
    result = {"valid": True, "errors": [], "warnings": [], "metrics": {}}
    errors: list[str] = result["errors"]  # type: ignore[assignment]
    warnings: list[str] = result["warnings"]  # type: ignore[assignment]
    workbook_path = workbook_path.expanduser()
    if not workbook_path.exists():
        errors.append(f"Workbook does not exist: {workbook_path}")
        result["valid"] = False
        return result

    model = None
    if trip_model_path:
        with trip_model_path.expanduser().open("r", encoding="utf-8") as handle:
            model = normalize_trip_model(json.load(handle))

    wb = load_workbook(workbook_path, data_only=False)
    expected_order = workbook_sheet_order(model) if model else wb.sheetnames
    if wb.sheetnames != expected_order:
        errors.append(f"Unexpected sheet order: {wb.sheetnames} != {expected_order}")

    specs = runtime_sheet_specs(model) if model else {name: {"widths": WIDTH_SPECS.get("Options", {}), "merges": []} for name in wb.sheetnames}
    for sheet_name in expected_order:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing expected sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        if ws.sheet_view.showGridLines:
            errors.append(f"{sheet_name}: gridlines should be hidden")
        if ws.freeze_panes not in (None, ""):
            errors.append(f"{sheet_name}: frozen panes should be absent")

    for sheet_name, spec in specs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for column_letter, expected_width in spec["widths"].items():
            actual_width = ws.column_dimensions[column_letter].width
            if actual_width is None or abs(actual_width - expected_width) > WIDTH_TOLERANCE:
                errors.append(f"{sheet_name}: unexpected width for {column_letter}: {actual_width} != {expected_width}")
        missing_merges = sorted(set(spec["merges"]) - {str(item) for item in ws.merged_cells.ranges})
        if missing_merges:
            errors.append(f"{sheet_name}: missing merged ranges {missing_merges}")

    check_headers(wb, model, errors)
    check_formulas(wb, model, errors)
    check_validations(wb, model, errors)
    check_daily_plan_pattern(wb["Daily Plan"], model, errors)

    result["valid"] = not errors
    result["metrics"] = {
        "sheet_count": len(wb.sheetnames),
        "daily_rows": wb["Daily Plan"].max_row - 3 if "Daily Plan" in wb.sheetnames else 0,
        "options_sheet": model["options_sheet_name"] if model else None,
        "primary_destination": destination_display_name(model) if model else None,
    }
    if not model:
        warnings.append("Validated without a trip model. Dynamic naming and planner-window length were not cross-checked.")
    return result


def check_headers(wb, model, errors: list[str]) -> None:
    header_checks = {
        "Bookings": ("A3:M3", HEADERS["Bookings"]),
        "Daily Plan": ("A3:L3", HEADERS["Daily Plan"]),
        model["options_sheet_name"] if model else wb.sheetnames[3]: ("A3:I3", HEADERS["Options"]),
        "Prep & Compliance": ("A3:I3", HEADERS["Prep & Compliance"]),
        "Pack List": ("A3:H3", HEADERS["Pack List"]),
        "Buy List": ("A3:J3", HEADERS["Buy List"]),
        "Sources": ("A3:F3", HEADERS["Sources"]),
    }
    for sheet_name, (_, expected_headers) in header_checks.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        actual_headers = [ws.cell(3, column).value for column in range(1, len(expected_headers) + 1)]
        if actual_headers != expected_headers:
            errors.append(f"{sheet_name}: unexpected headers {actual_headers}")

    overview = wb["Overview"]
    if overview["A4"].value != "Planner controls":
        errors.append("Overview: missing Planner controls section")
    if overview["E4"].value != "Trip status snapshot":
        errors.append("Overview: missing Trip status snapshot section")
    if overview["A22"].value != "Fixed anchors and protected moments":
        errors.append("Overview: missing fixed anchors section")


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def check_formulas(wb, model, errors: list[str]) -> None:
    overview = wb["Overview"]
    for coordinate in ("B10", "B11", "G5", "G6", "G7", "G8", "G9"):
        if not is_formula(overview[coordinate].value):
            errors.append(f"Overview: expected formula at {coordinate}")
    if not is_formula(overview["B19"].value):
        errors.append("Overview: expected formula at B19")

    bookings = wb["Bookings"]
    if bookings.max_row < 4:
        errors.append("Bookings: expected at least one data row")
        return
    for row in range(4, bookings.max_row + 1):
        if not is_formula(bookings[f"H{row}"].value):
            errors.append(f"Bookings: expected formula at H{row}")
        if not is_formula(bookings[f"I{row}"].value):
            errors.append(f"Bookings: expected formula at I{row}")
        if not is_formula(bookings[f"K{row}"].value):
            errors.append(f"Bookings: expected formula at K{row}")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and any(token in cell.value for token in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")):
                    errors.append(f"{sheet.title}: broken formula text at {cell.coordinate}: {cell.value}")


def validation_formulae(ws) -> list[str]:
    return sorted(str(validation.formula1) for validation in ws.data_validations.dataValidation)


def check_validations(wb, model, errors: list[str]) -> None:
    bookings_validations = validation_formulae(wb["Bookings"])
    if '"Booked,Planned,To confirm"' not in bookings_validations:
        errors.append("Bookings: missing status validation")

    daily_validations = validation_formulae(wb["Daily Plan"])
    if '"Locked,Flexible,Weather watch,Done"' not in daily_validations:
        errors.append("Daily Plan: missing status validation")

    prep_validations = validation_formulae(wb["Prep & Compliance"])
    if '"High,Medium,Low"' not in prep_validations or '"Not started,In progress,Review,Done"' not in prep_validations:
        errors.append("Prep & Compliance: missing required validations")

    pack_validations = validation_formulae(wb["Pack List"])
    if '"High,Medium,Low"' not in pack_validations or '"To pack,Need to buy,Packed,Clarify"' not in pack_validations:
        errors.append("Pack List: missing required validations")

    buy_validations = validation_formulae(wb["Buy List"])
    expected_buy_status = f'"Need to buy,Watch price,{model["buy_in_destination_label"]},Bought,Skip,Clarify"' if model else None
    if '"High,Medium,Low"' not in buy_validations:
        errors.append("Buy List: missing priority validation")
    if expected_buy_status and expected_buy_status not in buy_validations:
        errors.append(f"Buy List: missing expected status validation {expected_buy_status}")


def check_daily_plan_pattern(ws, model, errors: list[str]) -> None:
    data_rows = ws.max_row - 3
    if data_rows <= 0:
        errors.append("Daily Plan: no data rows found")
        return
    if data_rows % 3 != 0:
        errors.append(f"Daily Plan: row count after header is not a multiple of 3 ({data_rows})")
    expected_days = ((model["planner_end_date"] - model["planner_start_date"]).days + 1) if model else None
    if expected_days is not None and data_rows != expected_days * 3:
        errors.append(f"Daily Plan: expected {expected_days * 3} data rows, found {data_rows}")

    for row in range(4, ws.max_row + 1, 3):
        blocks = [ws[f"B{row + offset}"].value for offset in range(3) if row + offset <= ws.max_row]
        if blocks != list(BLOCKS):
            errors.append(f"Daily Plan: expected block sequence {BLOCKS} starting at row {row}, found {blocks}")
        dates = [ws[f"A{row + offset}"].value for offset in range(3) if row + offset <= ws.max_row]
        if len(set(dates)) > 1:
            errors.append(f"Daily Plan: expected same date across rows {row}-{row + 2}, found {dates}")


def main() -> None:
    args = parse_args()
    result = validate_workbook(Path(args.workbook), Path(args.trip_model) if args.trip_model else None)
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(json.dumps(result, indent=2))
    raise SystemExit(0 if result["valid"] else 1)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # pragma: no cover - CLI guard
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
