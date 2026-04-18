#!/usr/bin/env python3
"""Build a travel itinerary workbook from a canonical trip model JSON."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent

DATE_FORMAT_LONG = "ddd d mmm yyyy"
DATE_FORMAT_SHORT = "ddd d mmm"
BLOCKS = ("AM", "PM", "Eve")


def load_palette() -> dict[str, str]:
    path = SKILL_DIR / "assets" / "palette.json"
    with path.open("r", encoding="utf-8") as handle:
        raw = json.load(handle)
    return {key: value.replace("#", "") for key, value in raw.items()}


PALETTE = load_palette()

THIN_SIDE = Side(style="thin", color="CBD5E1")
MEDIUM_SIDE = Side(style="medium", color="94A3B8")


def solid_fill(name: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=PALETTE[name])


def calibri(size: float, *, bold: bool = False, color_name: str) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=PALETTE[color_name])


FILLS = {
    "navy": solid_fill("header_navy"),
    "teal": solid_fill("header_teal"),
    "static": solid_fill("static_label_fill"),
    "zebra_a": solid_fill("zebra_fill_a"),
    "zebra_b": solid_fill("zebra_fill_b"),
    "control": solid_fill("control_fill"),
    "metric": solid_fill("metric_fill"),
    "soft_caution": solid_fill("soft_caution_fill"),
    "caution_chip": solid_fill("caution_chip_fill"),
    "soft_error": solid_fill("soft_error_fill"),
    "success": solid_fill("success_fill"),
    "info": solid_fill("info_fill"),
    "buy_in_destination": solid_fill("buy_in_destination_fill"),
    "recovery": solid_fill("recovery_fill"),
}

FONTS = {
    "title": calibri(16, bold=True, color_name="text_white"),
    "subtitle": calibri(10, color_name="text_white"),
    "section_header": calibri(11, bold=True, color_name="text_white"),
    "column_header": calibri(10, bold=True, color_name="text_white"),
    "static_label": calibri(10, bold=True, color_name="text_static_grey"),
    "static_text": calibri(10, color_name="text_static_grey"),
    "editable": calibri(10, color_name="text_editable_blue"),
    "editable_bold": calibri(10, bold=True, color_name="text_editable_blue"),
    "imported": calibri(10, color_name="text_imported_green"),
    "imported_bold": calibri(10, bold=True, color_name="text_imported_green"),
    "formula": calibri(10, color_name="text_formula_black"),
    "minor": calibri(9, color_name="text_minor_slate"),
    "caution": calibri(10, color_name="text_caution_orange"),
    "caution_bold": calibri(10, bold=True, color_name="text_caution_orange"),
    "error_bold": calibri(10, bold=True, color_name="text_error_red"),
    "success_bold": calibri(10, bold=True, color_name="text_imported_green"),
}

ALIGNMENTS = {
    "left_center": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "left_top": Alignment(horizontal="left", vertical="top", wrap_text=True),
    "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "center_top": Alignment(horizontal="center", vertical="top", wrap_text=True),
}

BORDERS = {
    "none": Border(),
    "row": Border(bottom=THIN_SIDE),
    "section": Border(bottom=MEDIUM_SIDE),
    "column_header": Border(top=MEDIUM_SIDE, bottom=THIN_SIDE),
}

WIDTH_SPECS = {
    "Overview": {"A": 19, "B": 16, "C": 15, "D": 18, "E": 20, "F": 20, "G": 17, "H": 17, "I": 24},
    "Bookings": {
        "A": 13,
        "B": 32,
        "C": 14,
        "D": 18,
        "E": 24,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 8,
        "J": 16,
        "K": 15,
        "L": 40,
        "M": 46,
    },
    "Daily Plan": {
        "A": 12,
        "B": 8,
        "C": 14,
        "D": 34,
        "E": 30,
        "F": 30,
        "G": 18,
        "H": 24,
        "I": 20,
        "J": 10,
        "K": 14,
        "L": 28,
    },
    "Options": {"A": 16, "B": 26, "C": 24, "D": 10, "E": 16, "F": 13, "G": 12, "H": 40, "I": 16},
    "Prep & Compliance": {
        "A": 12,
        "B": 13,
        "C": 28,
        "D": 38,
        "E": 12,
        "F": 16,
        "G": 14,
        "H": 34,
        "I": 26,
    },
    "Pack List": {"A": 12, "B": 28, "C": 14, "D": 11, "E": 12, "F": 16, "G": 16, "H": 30},
    "Buy List": {
        "A": 13,
        "B": 12,
        "C": 14,
        "D": 26,
        "E": 24,
        "F": 12,
        "G": 17,
        "H": 10,
        "I": 34,
        "J": 24,
    },
    "Sources": {"A": 12, "B": 24, "C": 48, "D": 28, "E": 18, "F": 26},
}

MERGE_SPECS = {
    "Overview": [
        "A1:I1",
        "A2:I2",
        "A4:C4",
        "E4:I4",
        "G5:H5",
        "G6:H6",
        "G7:H7",
        "G8:H8",
        "G9:H9",
        "E10:I11",
        "A13:D13",
        "F13:I13",
        "F14:I14",
        "F15:I15",
        "F16:I16",
        "F17:I17",
        "F18:I18",
        "C19:D19",
        "A22:I22",
        "B23:D23",
        "E23:F23",
        "G23:I23",
        "B24:D24",
        "E24:F24",
        "G24:I24",
        "B25:D25",
        "E25:F25",
        "G25:I25",
        "B26:D26",
        "E26:F26",
        "G26:I26",
        "B27:D27",
        "E27:F27",
        "G27:I27",
        "B28:D28",
        "E28:F28",
        "G28:I28",
        "B29:D29",
        "E29:F29",
        "G29:I29",
        "B30:D30",
        "E30:F30",
        "G30:I30",
        "A33:I33",
        "B34:I34",
        "B35:I35",
        "B36:I36",
        "B37:I37",
        "B38:I38",
    ],
    "Bookings": ["A1:M1", "A2:M2"],
    "Daily Plan": ["A1:L1", "A2:L2"],
    "Options": ["A1:I1", "A2:I2"],
    "Prep & Compliance": ["A1:I1", "A2:I2"],
    "Pack List": ["A1:H1", "A2:H2"],
    "Buy List": ["A1:J1", "A2:J2"],
    "Sources": ["A1:F1", "A2:F2"],
}

HEADERS = {
    "Bookings": [
        "Type",
        "Item",
        "Date",
        "Time / local",
        "Area / route",
        "Booking ref",
        "Covered adults",
        "Planned adults",
        "Gap",
        "Status",
        "Coverage flag",
        "Action needed",
        "Notes / source",
    ],
    "Daily Plan": [
        "Date",
        "Block",
        "Phase",
        "Primary plan",
        "Fallback / option 1",
        "Fallback / split plan",
        "Area / base",
        "Shopping focus",
        "Fixed booking / time",
        "Pace",
        "Status",
        "Notes",
    ],
    "Options": [
        "Area",
        "Place / cluster",
        "Best for",
        "Indoor?",
        "Shopping strength",
        "Group fit",
        "Best day",
        "Why it earns a place",
        "Trigger / source",
    ],
    "Prep & Compliance": [
        "Due date",
        "Category",
        "Task",
        "Practical note",
        "Priority",
        "Status",
        "Owner",
        "Source URL",
        "Notes",
    ],
    "Pack List": ["Zone", "Item", "Who", "Need first?", "Priority", "Status", "Pack note", "Notes"],
    "Buy List": [
        "Phase",
        "Requester",
        "Category",
        "Item",
        "Best place / timing",
        "Priority",
        "Status",
        "Budget",
        "Why it matters",
        "Notes",
    ],
    "Sources": ["Type", "Source name", "URL or file note", "Used for", "Last checked / file date", "Notes"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--trip-model", required=True, help="Path to the canonical trip model JSON.")
    parser.add_argument(
        "--output",
        help="Output path or directory. If omitted, /mnt/data is preferred when present, otherwise the working directory.",
    )
    parser.add_argument("--normalized-model-out", help="Optional path to save the normalized trip model JSON.")
    parser.add_argument("--quiet", action="store_true", help="Suppress JSON summary output.")
    return parser.parse_args()


def parse_date(value: Any, field_name: str) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if value is None:
        raise ValueError(f"Missing required date field: {field_name}")
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f"Invalid ISO date for {field_name}: {value}") from exc


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def bullet_text(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    return value if value.startswith("• ") else f"• {value}"


def safe_filename_part(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    value = re.sub(r"[^\w\s-]", "", value, flags=re.UNICODE)
    value = re.sub(r"[\s-]+", "_", value).strip("_")
    return value


def format_date_span(start_date: date, end_date: date) -> str:
    if start_date.year == end_date.year and start_date.month == end_date.month:
        return f"{start_date.day} to {end_date.day} {end_date.strftime('%b %Y')}"
    if start_date.year == end_date.year:
        return f"{start_date.strftime('%d %b')} to {end_date.strftime('%d %b %Y')}"
    return f"{start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"


def destination_display_name(model: dict[str, Any]) -> str:
    return clean_text(model.get("primary_destination")) or "Travel"


def derive_options_sheet_name(model: dict[str, Any]) -> str:
    primary = destination_display_name(model)
    destinations = model.get("destinations", [])
    if primary.lower() == "tokyo":
        return "Tokyo Options"
    if primary and len(destinations) <= 1:
        return f"{primary} Options"
    return "Destination Options"


def derive_buy_in_destination_label(model: dict[str, Any]) -> str:
    primary = destination_display_name(model)
    if primary.lower() == "tokyo":
        return "Buy in Tokyo"
    if primary != "Travel":
        return f"Buy in {primary}"
    return "Buy in destination"


def workbook_sheet_order(model: dict[str, Any]) -> list[str]:
    return [
        "Overview",
        "Bookings",
        "Daily Plan",
        derive_options_sheet_name(model),
        "Prep & Compliance",
        "Pack List",
        "Buy List",
        "Sources",
    ]


def runtime_sheet_specs(model: dict[str, Any]) -> dict[str, dict[str, Any]]:
    options_name = derive_options_sheet_name(model)
    return {
        "Overview": {"widths": WIDTH_SPECS["Overview"], "merges": MERGE_SPECS["Overview"]},
        "Bookings": {"widths": WIDTH_SPECS["Bookings"], "merges": MERGE_SPECS["Bookings"]},
        "Daily Plan": {"widths": WIDTH_SPECS["Daily Plan"], "merges": MERGE_SPECS["Daily Plan"]},
        options_name: {"widths": WIDTH_SPECS["Options"], "merges": MERGE_SPECS["Options"]},
        "Prep & Compliance": {"widths": WIDTH_SPECS["Prep & Compliance"], "merges": MERGE_SPECS["Prep & Compliance"]},
        "Pack List": {"widths": WIDTH_SPECS["Pack List"], "merges": MERGE_SPECS["Pack List"]},
        "Buy List": {"widths": WIDTH_SPECS["Buy List"], "merges": MERGE_SPECS["Buy List"]},
        "Sources": {"widths": WIDTH_SPECS["Sources"], "merges": MERGE_SPECS["Sources"]},
    }


def normalize_review_flags(values: Iterable[Any]) -> list[str]:
    flags: list[str] = []
    for value in values:
        if isinstance(value, dict):
            text = clean_text(value.get("message"))
        else:
            text = clean_text(value)
        if text:
            flags.append(text)
    return flags


def normalize_travellers(values: Iterable[Any]) -> list[dict[str, Any]]:
    travellers: list[dict[str, Any]] = []
    for index, value in enumerate(values, start=1):
        if isinstance(value, str):
            travellers.append(
                {
                    "name": value,
                    "role": "Traveller",
                    "included_in_day_plans": True,
                    "notes": "",
                }
            )
            continue
        travellers.append(
            {
                "name": clean_text(value.get("name")) or f"Traveller {index}",
                "role": clean_text(value.get("role")) or "Traveller",
                "included_in_day_plans": bool(value.get("included_in_day_plans", True)),
                "notes": clean_text(value.get("notes")),
            }
        )
    return travellers


def normalize_bookings(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    bookings: list[dict[str, Any]] = []
    for value in values:
        booking = {
            "type": clean_text(value.get("type")) or "Booking",
            "item": clean_text(value.get("item")),
            "date": parse_date(value.get("date"), "bookings[].date"),
            "time_local": clean_text(value.get("time_local")),
            "area_route": clean_text(value.get("area_route")),
            "booking_ref": clean_text(value.get("booking_ref")),
            "covered_adults": int(value.get("covered_adults", 0) or 0),
            "status": clean_text(value.get("status")) or "Booked",
            "action_needed": clean_text(value.get("action_needed")),
            "notes_source": clean_text(value.get("notes_source")),
            "why_it_matters": clean_text(value.get("why_it_matters")),
        }
        if booking["item"]:
            bookings.append(booking)
    return bookings


def normalize_fixed_events(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    for value in values:
        anchor = clean_text(value.get("anchor"))
        if not anchor:
            continue
        events.append(
            {
                "date": parse_date(value.get("date"), "fixed_events[].date"),
                "anchor": anchor,
                "area_route": clean_text(value.get("area_route")),
                "why_it_matters": clean_text(value.get("why_it_matters")),
            }
        )
    return events


def normalize_daily_rows(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        block = clean_text(value.get("block")) or "AM"
        block = "Eve" if block.lower() == "eve" else block.upper() if block.lower() in {"am", "pm"} else block
        rows.append(
            {
                "date": parse_date(value.get("date"), "daily_rows[].date"),
                "block": block,
                "phase": clean_text(value.get("phase")),
                "primary_plan": clean_text(value.get("primary_plan")),
                "fallback_option_1": clean_text(value.get("fallback_option_1")),
                "fallback_split_plan": clean_text(value.get("fallback_split_plan")),
                "area_base": clean_text(value.get("area_base")),
                "shopping_focus": clean_text(value.get("shopping_focus")),
                "fixed_booking_time": clean_text(value.get("fixed_booking_time")),
                "pace": clean_text(value.get("pace")) or "Moderate",
                "status": clean_text(value.get("status")) or "Flexible",
                "notes": clean_text(value.get("notes")),
            }
        )
    return rows


def normalize_options_bank(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        rows.append(
            {
                "area": clean_text(value.get("area")),
                "place_cluster": clean_text(value.get("place_cluster")),
                "best_for": clean_text(value.get("best_for")),
                "indoor": clean_text(value.get("indoor")),
                "shopping_strength": int(value.get("shopping_strength", 0) or 0),
                "group_fit": clean_text(value.get("group_fit")),
                "best_day": clean_text(value.get("best_day")),
                "why_it_earns_a_place": clean_text(value.get("why_it_earns_a_place")),
                "trigger_source": clean_text(value.get("trigger_source")),
            }
        )
    return [row for row in rows if row["place_cluster"]]


def normalize_prep_tasks(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        due_raw = value.get("due_date")
        rows.append(
            {
                "due_date": parse_date(due_raw, "prep_tasks[].due_date") if due_raw else None,
                "category": clean_text(value.get("category")),
                "task": clean_text(value.get("task")),
                "practical_note": clean_text(value.get("practical_note")),
                "priority": clean_text(value.get("priority")) or "Medium",
                "status": clean_text(value.get("status")) or "Not started",
                "owner": clean_text(value.get("owner")) or "Planner",
                "source_url": clean_text(value.get("source_url")),
                "notes": clean_text(value.get("notes")),
            }
        )
    return [row for row in rows if row["task"]]


def normalize_pack_items(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        rows.append(
            {
                "zone": clean_text(value.get("zone")),
                "item": clean_text(value.get("item")),
                "who": clean_text(value.get("who")),
                "need_first": clean_text(value.get("need_first")) or "No",
                "priority": clean_text(value.get("priority")) or "Medium",
                "status": clean_text(value.get("status")) or "To pack",
                "pack_note": clean_text(value.get("pack_note")),
                "notes": clean_text(value.get("notes")),
            }
        )
    return [row for row in rows if row["item"]]


def normalize_buy_items(values: Iterable[dict[str, Any]], buy_in_destination_label: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        status = clean_text(value.get("status")) or "Need to buy"
        if status.lower() == "buy in destination":
            status = buy_in_destination_label
        rows.append(
            {
                "phase": clean_text(value.get("phase")),
                "requester": clean_text(value.get("requester")),
                "category": clean_text(value.get("category")),
                "item": clean_text(value.get("item")),
                "best_place_timing": clean_text(value.get("best_place_timing")),
                "priority": clean_text(value.get("priority")) or "Medium",
                "status": status,
                "budget": clean_text(value.get("budget")),
                "why_it_matters": clean_text(value.get("why_it_matters")),
                "notes": clean_text(value.get("notes")),
            }
        )
    return [row for row in rows if row["item"]]


def normalize_sources(values: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for value in values:
        rows.append(
            {
                "type": clean_text(value.get("type")),
                "source_name": clean_text(value.get("source_name")),
                "url_or_file_note": clean_text(value.get("url_or_file_note")),
                "used_for": clean_text(value.get("used_for")),
                "last_checked_or_file_date": clean_text(value.get("last_checked_or_file_date")),
                "notes": clean_text(value.get("notes")),
            }
        )
    return [row for row in rows if row["source_name"]]


def synthesize_bookings(raw: dict[str, Any]) -> list[dict[str, Any]]:
    synthesized: list[dict[str, Any]] = []
    for stay in clean_list(raw.get("accommodation")):
        name = clean_text(stay.get("name"))
        if not name or not stay.get("check_in_date"):
            continue
        synthesized.append(
            {
                "type": "Hotel",
                "item": name,
                "date": stay.get("check_in_date"),
                "time_local": clean_text(stay.get("time_local")) or "Check-in / check-out",
                "area_route": clean_text(stay.get("area")),
                "booking_ref": clean_text(stay.get("booking_ref")),
                "covered_adults": stay.get("covered_adults", 0),
                "status": "Booked",
                "action_needed": clean_text(stay.get("notes")),
                "notes_source": clean_text(stay.get("notes")),
                "why_it_matters": clean_text(stay.get("why_it_matters")),
            }
        )
    for leg in clean_list(raw.get("transport_legs")):
        route = clean_text(leg.get("route"))
        if not route or not leg.get("date"):
            continue
        synthesized.append(
            {
                "type": clean_text(leg.get("type")) or "Transport",
                "item": route,
                "date": leg.get("date"),
                "time_local": clean_text(leg.get("time_local")),
                "area_route": route,
                "booking_ref": clean_text(leg.get("booking_ref")),
                "covered_adults": leg.get("covered_adults", 0),
                "status": "Booked",
                "action_needed": clean_text(leg.get("notes")),
                "notes_source": clean_text(leg.get("notes")),
                "why_it_matters": clean_text(leg.get("notes")),
            }
        )
    return synthesized


def destination_phase_label(model: dict[str, Any]) -> str:
    primary = destination_display_name(model)
    return primary if primary != "Travel" else "In-destination"


def is_conference_like(text: str) -> bool:
    lowered = text.lower()
    return "conference" in lowered or "work" in lowered


def default_phase_for_date(model: dict[str, Any], current_date: date) -> str:
    if current_date < model["departure_date"]:
        return "Pre-departure"
    if current_date < model["stay_start_date"]:
        return "Travel"
    if model["stay_start_date"] <= current_date <= model["stay_end_date"]:
        for event in model["fixed_events"]:
            if event["date"] == current_date and is_conference_like(event["anchor"]):
                return "Conference"
        return destination_phase_label(model)
    if current_date <= model["final_home_arrival_date"]:
        return "Return"
    return "Recovery"


def default_daily_row(model: dict[str, Any], current_date: date, block: str) -> dict[str, Any]:
    phase = default_phase_for_date(model, current_date)
    if phase == "Pre-departure":
        primary_plan = "Keep this block for final prep, packing, and confirming the essentials."
        pace = "Easy"
    elif phase == "Travel":
        primary_plan = "Travel or transfer block. Keep the logistics simple and the expectations light."
        pace = "Moderate"
    elif phase == "Conference":
        primary_plan = "Protect the fixed work commitment and keep regroup options nearby."
        pace = "Structured"
    elif phase == "Return":
        primary_plan = "Keep the return sequence light and leave buffer for airport or transit delays."
        pace = "Easy"
    elif phase == "Recovery":
        primary_plan = "Use this block for recovery, unpacking, and a gentle reset."
        pace = "Easy"
    else:
        primary_plan = "Open planning block. Use this slot for the strongest nearby cluster."
        pace = "Moderate"

    return {
        "date": current_date,
        "block": block,
        "phase": phase,
        "primary_plan": primary_plan,
        "fallback_option_1": "",
        "fallback_split_plan": "",
        "area_base": "",
        "shopping_focus": "",
        "fixed_booking_time": "",
        "pace": pace,
        "status": "Locked" if phase in {"Conference"} else "Flexible",
        "notes": "",
    }


def build_daily_schedule(model: dict[str, Any]) -> list[dict[str, Any]]:
    keyed_rows = {(row["date"], row["block"]): row for row in model["daily_rows"]}
    rows: list[dict[str, Any]] = []
    current_date = model["planner_start_date"]
    while current_date <= model["planner_end_date"]:
        for block in BLOCKS:
            base = default_daily_row(model, current_date, block)
            row = dict(base)
            provided = keyed_rows.get((current_date, block))
            if provided:
                for key, value in provided.items():
                    if value not in {"", None}:
                        row[key] = value
            rows.append(row)
        current_date += timedelta(days=1)
    return rows


def derive_fixed_anchors(model: dict[str, Any]) -> list[dict[str, Any]]:
    anchors: list[dict[str, Any]] = []
    for booking in model["bookings"]:
        anchors.append(
            {
                "date": booking["date"],
                "anchor": booking["item"],
                "area_route": booking["area_route"],
                "why_it_matters": booking["why_it_matters"] or booking["action_needed"] or "Confirmed anchor.",
            }
        )
    anchors.extend(model["fixed_events"])
    anchors.sort(key=lambda item: (item["date"], item["anchor"]))
    return anchors


def build_review_lines(model: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for flag in model["review_flags"]:
        if flag not in lines:
            lines.append(flag)
    for note in model["notes_normalised"]:
        text = clean_text(note)
        if text and text not in lines:
            lines.append(text)
    if not lines:
        lines.append("No material booking or note-normalization gaps are open right now.")
    return lines


def summary_cleanup_note(model: dict[str, Any]) -> str:
    if model["review_flags"]:
        top = model["review_flags"][:3]
        joined = "; ".join(clean_text(flag).rstrip(".") for flag in top)
        return f"Main clean-up jobs: {joined}."
    if model["assumptions"]:
        joined = "; ".join(clean_text(item).rstrip(".") for item in model["assumptions"][:3])
        return f"Main clean-up jobs: review these assumptions before travel: {joined}."
    return "Main clean-up jobs: none. Keep Bookings, Prep & Compliance, and Sources current as the trip evolves."


def overview_title(model: dict[str, Any]) -> str:
    prefix = clean_text(model.get("trip_title")) or f"{destination_display_name(model)} travel planner"
    return f"{prefix} | {format_date_span(model['planner_start_date'], model['planner_end_date'])}"


def overview_subtitle(model: dict[str, Any]) -> str:
    names = [traveller["name"] for traveller in model["travellers"][:4]]
    if len(model["travellers"]) > 4:
        names[-1] = f"{names[-1]} + {len(model['travellers']) - 3} more"
    party_text = ", ".join(name for name in names if name)
    split_needed = any("split" in row["fallback_split_plan"].lower() for row in model["daily_rows"]) or any(
        "coverage" in flag.lower() or "ticket" in flag.lower() for flag in model["review_flags"]
    )
    shopping_text = "Shopping stays visible" if model["shopping_objectives"] or model["buy_items"] else "The workbook stays tidy"
    anchor_text = "fixed commitments stay protected"
    split_text = "split fallbacks are baked in" if split_needed else "fallbacks stay visible"
    if len(model["destinations"]) > 1:
        base_text = "Multi-base travel plan."
    else:
        base_text = f"{destination_display_name(model)}-led base plan."
    if party_text:
        return f"{base_text} Built for {party_text}. {shopping_text}, {anchor_text}, and {split_text}."
    return f"{base_text} {shopping_text}, {anchor_text}, and {split_text}."


def choose_output_path(model: dict[str, Any], output_arg: str | None) -> Path:
    filename = build_default_filename(model)
    if output_arg:
        candidate = Path(output_arg).expanduser()
        if candidate.suffix.lower() == ".xlsx":
            return candidate
        return candidate / filename
    if Path("/mnt/data").exists():
        return Path("/mnt/data") / filename
    return Path.cwd() / filename


def build_default_filename(model: dict[str, Any]) -> str:
    month = model["stay_start_date"].strftime("%B")
    year = model["stay_start_date"].strftime("%Y")
    primary = destination_display_name(model)
    if primary != "Travel":
        prefix = safe_filename_part(primary) or "Travel"
        return f"{prefix}_Travel_Itinerary_{month}_{year}.xlsx"
    return f"Travel_Itinerary_{month}_{year}.xlsx"


def normalize_trip_model(raw: dict[str, Any]) -> dict[str, Any]:
    model: dict[str, Any] = {}
    model["trip_title"] = clean_text(raw.get("trip_title")) or f"{clean_text(raw.get('primary_destination'))} travel planner"
    model["primary_destination"] = clean_text(raw.get("primary_destination"))
    model["destinations"] = []
    for destination in clean_list(raw.get("destinations")):
        if isinstance(destination, str):
            model["destinations"].append({"name": destination, "country": "", "is_primary": False, "base": False})
        else:
            name = clean_text(destination.get("name"))
            if name:
                model["destinations"].append(
                    {
                        "name": name,
                        "country": clean_text(destination.get("country")),
                        "is_primary": bool(destination.get("is_primary", False)),
                        "base": bool(destination.get("base", False)),
                    }
                )
    if not model["destinations"] and model["primary_destination"]:
        model["destinations"].append({"name": model["primary_destination"], "country": "", "is_primary": True, "base": True})

    model["travellers"] = normalize_travellers(clean_list(raw.get("travellers")))
    model["stay_start_date"] = parse_date(raw.get("stay_start_date"), "stay_start_date")
    model["stay_end_date"] = parse_date(raw.get("stay_end_date"), "stay_end_date")
    model["departure_date"] = parse_date(raw.get("departure_date"), "departure_date")
    model["final_home_arrival_date"] = parse_date(raw.get("final_home_arrival_date"), "final_home_arrival_date")
    model["planner_start_date"] = parse_date(raw.get("planner_start_date"), "planner_start_date") if raw.get("planner_start_date") else model["departure_date"] - timedelta(days=1)
    model["planner_end_date"] = parse_date(raw.get("planner_end_date"), "planner_end_date") if raw.get("planner_end_date") else model["final_home_arrival_date"] + timedelta(days=1)
    model["transport_legs"] = clean_list(raw.get("transport_legs"))
    model["accommodation"] = clean_list(raw.get("accommodation"))
    model["bookings"] = normalize_bookings(clean_list(raw.get("bookings")) or synthesize_bookings(raw))
    model["fixed_events"] = normalize_fixed_events(clean_list(raw.get("fixed_events")))
    model["daily_rows"] = normalize_daily_rows(clean_list(raw.get("daily_rows")))
    model["options_bank"] = normalize_options_bank(clean_list(raw.get("options_bank")))
    model["prep_tasks"] = normalize_prep_tasks(clean_list(raw.get("prep_tasks")))
    model["pack_items"] = normalize_pack_items(clean_list(raw.get("pack_items")))
    model["buy_items"] = normalize_buy_items(clean_list(raw.get("buy_items")), derive_buy_in_destination_label(model))
    model["sources"] = normalize_sources(clean_list(raw.get("sources")))
    model["assumptions"] = [clean_text(value) for value in clean_list(raw.get("assumptions")) if clean_text(value)]
    model["review_flags"] = normalize_review_flags(clean_list(raw.get("review_flags")))
    model["timezones"] = [clean_text(value) for value in clean_list(raw.get("timezones")) if clean_text(value)]
    model["shopping_objectives"] = [item for item in clean_list(raw.get("shopping_objectives")) if isinstance(item, dict)]
    model["notes_raw"] = [clean_text(value) for value in clean_list(raw.get("notes_raw")) if clean_text(value)]
    model["notes_normalised"] = [clean_text(value) for value in clean_list(raw.get("notes_normalised")) if clean_text(value)]
    model["options_sheet_name"] = derive_options_sheet_name(model)
    model["buy_in_destination_label"] = derive_buy_in_destination_label(model)
    model["planned_adults"] = sum(1 for traveller in model["travellers"] if traveller.get("included_in_day_plans", True))
    model["fixed_anchor_rows"] = derive_fixed_anchors(model)
    model["daily_rows"] = build_daily_schedule(model)
    return model


def style_cell(
    cell,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def style_range(
    ws,
    cell_range: str,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            style_cell(cell, fill=fill, font=font, alignment=alignment, border=border, number_format=number_format)


def merge_value(
    ws,
    cell_range: str,
    value: Any,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    style_range(ws, cell_range, fill=fill, font=font, alignment=alignment, border=border, number_format=number_format)
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":", 1)[0]]
    top_left.value = value


def value_cell(
    ws,
    coordinate: str,
    value: Any,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    cell = ws[coordinate]
    cell.value = value
    style_cell(cell, fill=fill, font=font, alignment=alignment, border=border, number_format=number_format)


def apply_sheet_geometry(ws, widths: dict[str, float], merges: list[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def apply_row_fill(ws, row_index: int, end_column: int, fill: PatternFill) -> None:
    for column in range(1, end_column + 1):
        style_cell(ws.cell(row_index, column), fill=fill, border=BORDERS["row"])


def zebra_fill(row_number: int) -> PatternFill:
    return FILLS["zebra_a"] if row_number % 2 == 0 else FILLS["zebra_b"]


def phase_fill(model: dict[str, Any], phase: str) -> PatternFill:
    lowered = phase.lower()
    if "pre" in lowered:
        return FILLS["static"]
    if "travel" in lowered or "transit" in lowered:
        return FILLS["info"]
    if "conference" in lowered or "work" in lowered:
        return FILLS["buy_in_destination"]
    if "return" in lowered:
        return FILLS["caution_chip"]
    if "recovery" in lowered:
        return FILLS["recovery"]
    if clean_text(phase).lower() == destination_phase_label(model).lower():
        return FILLS["success"]
    return FILLS["success"]


def booking_status_style(status: str) -> tuple[PatternFill, Font]:
    lowered = status.lower()
    if lowered == "booked":
        return FILLS["success"], FONTS["success_bold"]
    if lowered == "planned":
        return FILLS["info"], FONTS["editable_bold"]
    return FILLS["caution_chip"], FONTS["caution_bold"]


def daily_status_style(status: str) -> tuple[PatternFill, Font]:
    lowered = status.lower()
    if lowered == "locked":
        return FILLS["success"], FONTS["editable_bold"]
    if lowered == "flexible":
        return FILLS["info"], FONTS["editable_bold"]
    if lowered == "weather watch":
        return FILLS["caution_chip"], FONTS["caution_bold"]
    return FILLS["success"], FONTS["minor"]


def prep_status_style(status: str) -> tuple[PatternFill, Font]:
    lowered = status.lower()
    if lowered == "not started":
        return FILLS["soft_error"], FONTS["error_bold"]
    if lowered == "in progress":
        return FILLS["info"], FONTS["editable_bold"]
    if lowered == "review":
        return FILLS["caution_chip"], FONTS["caution_bold"]
    return FILLS["success"], FONTS["success_bold"]


def pack_status_style(status: str) -> tuple[PatternFill, Font]:
    lowered = status.lower()
    if lowered == "packed":
        return FILLS["success"], FONTS["success_bold"]
    if lowered == "to pack":
        return FILLS["info"], FONTS["editable_bold"]
    return FILLS["caution_chip"], FONTS["caution_bold"]


def buy_status_style(status: str, buy_in_destination_label: str) -> tuple[PatternFill, Font]:
    lowered = status.lower()
    if lowered == "bought":
        return FILLS["success"], FONTS["success_bold"]
    if lowered == "watch price":
        return FILLS["info"], FONTS["editable_bold"]
    if lowered == buy_in_destination_label.lower():
        return FILLS["buy_in_destination"], FONTS["editable_bold"]
    if lowered == "skip":
        return FILLS["static"], FONTS["static_text"]
    return FILLS["caution_chip"], FONTS["caution_bold"]


def add_list_validation(ws, target_range: str, values: list[str]) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    validation.add(target_range)
    ws.add_data_validation(validation)


def ensure_minimum_rows(rows: list[dict[str, Any]], placeholder: dict[str, Any]) -> list[dict[str, Any]]:
    return rows if rows else [placeholder]


def build_overview(ws, model: dict[str, Any], ranges: dict[str, str]) -> None:
    widths = WIDTH_SPECS["Overview"]
    apply_sheet_geometry(ws, widths, MERGE_SPECS["Overview"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[10].height = 28
    ws.row_dimensions[11].height = 26
    for row_index in range(14, 19):
        ws.row_dimensions[row_index].height = 34 if row_index != 18 else 28
    for row_index in range(24, 31):
        ws.row_dimensions[row_index].height = 28

    merge_value(ws, "A1:I1", overview_title(model), fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:I2",
        overview_subtitle(model),
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )

    merge_value(ws, "A4:C4", "Planner controls", fill=FILLS["navy"], font=FONTS["section_header"], alignment=ALIGNMENTS["left_center"], border=BORDERS["section"])
    merge_value(ws, "E4:I4", "Trip status snapshot", fill=FILLS["navy"], font=FONTS["section_header"], alignment=ALIGNMENTS["left_center"], border=BORDERS["section"])

    start_label = f"{destination_display_name(model)} check-in" if destination_display_name(model) != "Travel" else "Stay start"
    end_label = f"{destination_display_name(model)} check-out" if destination_display_name(model) != "Travel" else "Stay end"
    control_rows = [
        ("A5", "B5", "Plan prepared on", date.today(), DATE_FORMAT_LONG),
        ("A6", "B6", "Planner start", model["planner_start_date"], DATE_FORMAT_LONG),
        ("A7", "B7", start_label, model["stay_start_date"], DATE_FORMAT_LONG),
        ("A8", "B8", end_label, model["stay_end_date"], DATE_FORMAT_LONG),
        ("A9", "B9", "Planner end", model["planner_end_date"], DATE_FORMAT_LONG),
        ("A10", "B10", "Hotel nights", "=B8-B7", None),
        ("A11", "B11", "Planner days", "=B9-B6+1", None),
    ]
    for label_coord, value_coord, label, value, number_format in control_rows:
        value_cell(ws, label_coord, label, fill=FILLS["static"], font=FONTS["static_label"], alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])
        font = FONTS["formula"] if isinstance(value, str) and value.startswith("=") else FONTS["editable"]
        value_cell(ws, value_coord, value, fill=FILLS["control"], font=font, alignment=ALIGNMENTS["center"], border=BORDERS["row"], number_format=number_format)

    snapshot_rows = [
        ("E5", "G5", "Anchors in workbook", f'=COUNTIF({ranges["bookings_item_range"]},"<>")', FILLS["metric"]),
        ("E6", "G6", "Coverage items to review", f'=COUNTIF({ranges["bookings_coverage_range"]},"Needs review")', FILLS["control"]),
        (
            "E7",
            "G7",
            "Open prep tasks",
            f'=COUNTIF({ranges["prep_status_range"]},"Not started")+COUNTIF({ranges["prep_status_range"]},"In progress")+COUNTIF({ranges["prep_status_range"]},"Review")',
            FILLS["metric"],
        ),
        (
            "E8",
            "G8",
            "Pack items still open",
            f'=COUNTIF({ranges["pack_status_range"]},"To pack")+COUNTIF({ranges["pack_status_range"]},"Need to buy")+COUNTIF({ranges["pack_status_range"]},"Clarify")',
            FILLS["control"],
        ),
        (
            "E9",
            "G9",
            "Buy items still open",
            f'=COUNTIF({ranges["buy_status_range"]},"Need to buy")+COUNTIF({ranges["buy_status_range"]},"Watch price")+COUNTIF({ranges["buy_status_range"]},"{model["buy_in_destination_label"]}")+COUNTIF({ranges["buy_status_range"]},"Clarify")',
            FILLS["metric"],
        ),
    ]
    for label_coord, value_coord, label, formula, fill in snapshot_rows:
        value_cell(ws, label_coord, label, fill=FILLS["static"], font=FONTS["static_label"], alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])
        row_number = value_coord[1:]
        merge_value(
            ws,
            f"G{row_number}:H{row_number}",
            formula,
            fill=fill,
            font=FONTS["formula"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["row"],
            number_format="0",
        )
    merge_value(
        ws,
        "E10:I11",
        summary_cleanup_note(model),
        fill=FILLS["soft_caution"],
        font=FONTS["caution"],
        alignment=ALIGNMENTS["left_center"],
        border=BORDERS["row"],
    )

    merge_value(ws, "A13:D13", "Travel party", fill=FILLS["navy"], font=FONTS["section_header"], alignment=ALIGNMENTS["left_center"], border=BORDERS["section"])
    merge_value(
        ws,
        "F13:I13",
        "Known booking gaps to confirm",
        fill=FILLS["navy"],
        font=FONTS["section_header"],
        alignment=ALIGNMENTS["left_center"],
        border=BORDERS["section"],
    )

    traveller_headers = ["Name", "Role", "Included in day plans?", "Notes"]
    for idx, header in enumerate(traveller_headers, start=1):
        value_cell(
            ws,
            f"{get_column_letter(idx)}14",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    travellers = model["travellers"][:4]
    overflow_count = max(0, len(model["travellers"]) - 4)
    for display_row, traveller in zip(range(15, 19), travellers):
        fill = zebra_fill(display_row)
        value_cell(ws, f"A{display_row}", traveller["name"], fill=fill, font=FONTS["editable"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"B{display_row}", traveller["role"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(
            ws,
            f"C{display_row}",
            "Yes" if traveller["included_in_day_plans"] else "No",
            fill=fill,
            font=FONTS["static_text"],
            alignment=ALIGNMENTS["left_top"],
            border=BORDERS["row"],
        )
        note_text = traveller["notes"]
        if overflow_count and display_row == 18:
            suffix = f" Plus {overflow_count} more traveller(s) in the canonical model."
            note_text = f"{note_text}{suffix}".strip()
        value_cell(ws, f"D{display_row}", note_text, fill=fill, font=FONTS["editable"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])

    for row_index in range(15 + len(travellers), 19):
        fill = zebra_fill(row_index)
        for column_letter in "ABCD":
            value_cell(ws, f"{column_letter}{row_index}", "", fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])

    review_lines = build_review_lines(model)
    review_slots = [14, 15, 16, 17, 18]
    for index, row_index in enumerate(review_slots):
        fill = FILLS["soft_caution"] if index < 4 else FILLS["static"]
        font = FONTS["caution"] if index < 4 else FONTS["minor"]
        text = bullet_text(review_lines[index]) if index < len(review_lines) else ""
        if index == 4 and len(review_lines) > 5:
            text = bullet_text(f"{review_lines[4]} Additional flags remain in Bookings and Prep & Compliance.")
        merge_value(ws, f"F{row_index}:I{row_index}", text, fill=fill, font=font, alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])

    value_cell(ws, "A19", "Planned adults", fill=FILLS["static"], font=FONTS["static_label"], alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])
    planned_formula = f"={model['planned_adults']}" if len(model["travellers"]) > 4 else '=COUNTIF(B15:B18,"<>")'
    value_cell(ws, "B19", planned_formula, fill=FILLS["metric"], font=FONTS["formula"], alignment=ALIGNMENTS["center"], border=BORDERS["row"], number_format="0")
    merge_value(
        ws,
        "C19:D19",
        "Edit names above if the group changes.",
        fill=FILLS["static"],
        font=FONTS["minor"],
        alignment=ALIGNMENTS["left_top"],
        border=BORDERS["row"],
    )

    merge_value(
        ws,
        "A22:I22",
        "Fixed anchors and protected moments",
        fill=FILLS["navy"],
        font=FONTS["section_header"],
        alignment=ALIGNMENTS["left_center"],
        border=BORDERS["section"],
    )
    value_cell(ws, "A23", "Date", fill=FILLS["teal"], font=FONTS["column_header"], alignment=ALIGNMENTS["center"], border=BORDERS["column_header"])
    merge_value(ws, "B23:D23", "Anchor", fill=FILLS["teal"], font=FONTS["column_header"], alignment=ALIGNMENTS["center"], border=BORDERS["column_header"])
    merge_value(ws, "E23:F23", "Area / route", fill=FILLS["teal"], font=FONTS["column_header"], alignment=ALIGNMENTS["center"], border=BORDERS["column_header"])
    merge_value(ws, "G23:I23", "Why it matters", fill=FILLS["teal"], font=FONTS["column_header"], alignment=ALIGNMENTS["center"], border=BORDERS["column_header"])

    anchor_rows = model["fixed_anchor_rows"][:7]
    if len(model["fixed_anchor_rows"]) > 7 and anchor_rows:
        anchor_rows[-1] = {
            "date": anchor_rows[-1]["date"],
            "anchor": anchor_rows[-1]["anchor"],
            "area_route": anchor_rows[-1]["area_route"],
            "why_it_matters": f'{anchor_rows[-1]["why_it_matters"]} More anchors exist in Bookings and Daily Plan.',
        }
    for row_index, anchor in zip(range(24, 31), anchor_rows):
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", anchor["date"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["center"], border=BORDERS["row"], number_format=DATE_FORMAT_SHORT)
        merge_value(ws, f"B{row_index}:D{row_index}", anchor["anchor"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        merge_value(ws, f"E{row_index}:F{row_index}", anchor["area_route"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        merge_value(ws, f"G{row_index}:I{row_index}", anchor["why_it_matters"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
    for row_index in range(24 + len(anchor_rows), 31):
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", "", fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        merge_value(ws, f"B{row_index}:D{row_index}", "", fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        merge_value(ws, f"E{row_index}:F{row_index}", "", fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        merge_value(ws, f"G{row_index}:I{row_index}", "", fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])

    merge_value(ws, "A33:I33", "Legend and edit notes", fill=FILLS["navy"], font=FONTS["section_header"], alignment=ALIGNMENTS["left_center"], border=BORDERS["section"])
    legend_rows = [
        ("Blue text", "Editable fields and status cells you will likely touch."),
        ("Green text", "Imported or confirmed details from files and sources."),
        ("Black text", "Formula outputs and workbook logic."),
        ("Orange fill", "Review, ticket mismatch, or clarify-before-you-fly items."),
        ("Purple / teal fill", "Controls, work phases, and headline counters."),
    ]
    for row_index, (label, note) in zip(range(34, 39), legend_rows):
        value_cell(ws, f"A{row_index}", label, fill=FILLS["static"], font=FONTS["static_label"], alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])
        merge_value(ws, f"B{row_index}:I{row_index}", note, fill=FILLS["static"], font=FONTS["minor"], alignment=ALIGNMENTS["left_center"], border=BORDERS["row"])


def build_bookings(ws, model: dict[str, Any], ranges: dict[str, str]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Bookings"], MERGE_SPECS["Bookings"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:M1", "Bookings and coverage check", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:M2",
        "This sheet keeps the confirmed anchors tidy and makes the booking gaps obvious.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Bookings"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    placeholder = {
        "type": "",
        "item": "",
        "date": None,
        "time_local": "",
        "area_route": "",
        "booking_ref": "",
        "covered_adults": 0,
        "status": "To confirm",
        "action_needed": "",
        "notes_source": "",
        "why_it_matters": "",
    }
    booking_rows = ensure_minimum_rows(model["bookings"], placeholder)
    end_row = 3 + len(booking_rows)
    add_list_validation(ws, f"J4:J{end_row}", ["Booked", "Planned", "To confirm"])

    for row_index, booking in enumerate(booking_rows, start=4):
        ws.row_dimensions[row_index].height = 36
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", booking["type"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        for column_letter, key in zip(("B", "C", "D", "E", "F", "G"), ("item", "date", "time_local", "area_route", "booking_ref", "covered_adults")):
            number_format = DATE_FORMAT_SHORT if key == "date" and booking["date"] else None
            alignment = ALIGNMENTS["center"] if key in {"date", "covered_adults"} else ALIGNMENTS["left_top"]
            if key == "covered_adults":
                font = FONTS["imported"]
            else:
                font = FONTS["imported"]
            value_cell(
                ws,
                f"{column_letter}{row_index}",
                booking[key],
                fill=fill,
                font=font,
                alignment=alignment,
                border=BORDERS["row"],
                number_format=number_format,
            )
        value_cell(
            ws,
            f"H{row_index}",
            "=Overview!$B$19",
            fill=FILLS["control"],
            font=FONTS["formula"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["row"],
            number_format="0",
        )
        value_cell(
            ws,
            f"I{row_index}",
            f"=IF(H{row_index}-G{row_index}>0,H{row_index}-G{row_index},0)",
            fill=FILLS["metric"],
            font=FONTS["formula"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["row"],
            number_format="0",
        )
        status_fill, status_font = booking_status_style(booking["status"])
        value_cell(ws, f"J{row_index}", booking["status"], fill=status_fill, font=status_font, alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(
            ws,
            f"K{row_index}",
            f'=IF(I{row_index}>0,"Needs review","OK")',
            fill=FILLS["soft_caution"],
            font=FONTS["formula"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["row"],
        )
        value_cell(
            ws,
            f"L{row_index}",
            booking["action_needed"],
            fill=FILLS["soft_caution"],
            font=FONTS["caution"],
            alignment=ALIGNMENTS["left_top"],
            border=BORDERS["row"],
        )
        value_cell(
            ws,
            f"M{row_index}",
            booking["notes_source"],
            fill=fill,
            font=FONTS["minor"],
            alignment=ALIGNMENTS["left_top"],
            border=BORDERS["row"],
        )


def build_daily_plan(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Daily Plan"], MERGE_SPECS["Daily Plan"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:L1", "Daily plan with primary plan + two fallbacks", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:L2",
        "Three rows per day keeps the trip flexible without losing the shape of the stay.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Daily Plan"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    end_row = 3 + len(model["daily_rows"])
    add_list_validation(ws, f"K4:K{end_row}", ["Locked", "Flexible", "Weather watch", "Done"])

    for row_index, row in enumerate(model["daily_rows"], start=4):
        ws.row_dimensions[row_index].height = 46
        fill = phase_fill(model, row["phase"])
        for column in range(1, 13):
            style_cell(ws.cell(row_index, column), fill=fill, border=BORDERS["row"])
        value_cell(ws, f"A{row_index}", row["date"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"], number_format=DATE_FORMAT_SHORT)
        value_cell(ws, f"B{row_index}", row["block"], fill=fill, font=FONTS["static_label"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"C{row_index}", row["phase"], fill=fill, font=FONTS["static_label"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", row["primary_plan"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", row["fallback_option_1"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"F{row_index}", row["fallback_split_plan"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"G{row_index}", row["area_base"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"H{row_index}", row["shopping_focus"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"I{row_index}", row["fixed_booking_time"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"J{row_index}", row["pace"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        status_fill, status_font = daily_status_style(row["status"])
        value_cell(ws, f"K{row_index}", row["status"], fill=status_fill, font=status_font, alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"L{row_index}", row["notes"], fill=fill, font=FONTS["minor"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def build_options_sheet(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Options"], MERGE_SPECS["Options"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:I1", f"{destination_display_name(model)} option bank" if destination_display_name(model) != "Travel" else "Destination option bank", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:I2",
        "Use this as the overflow parking lot for ideas so the daily sheet stays tidy.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Options"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    options = ensure_minimum_rows(
        model["options_bank"],
        {
            "area": "",
            "place_cluster": "",
            "best_for": "",
            "indoor": "",
            "shopping_strength": "",
            "group_fit": "",
            "best_day": "",
            "why_it_earns_a_place": "",
            "trigger_source": "",
        },
    )
    for row_index, item in enumerate(options, start=4):
        ws.row_dimensions[row_index].height = 38
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", item["area"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"B{row_index}", item["place_cluster"], fill=fill, font=FONTS["imported"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"C{row_index}", item["best_for"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", item["indoor"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", item["shopping_strength"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"F{row_index}", item["group_fit"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"G{row_index}", item["best_day"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"H{row_index}", item["why_it_earns_a_place"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"I{row_index}", item["trigger_source"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def build_prep_sheet(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Prep & Compliance"], MERGE_SPECS["Prep & Compliance"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:I1", "Prep, health, immigration and flight admin", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:I2",
        "This is the hard-stop sheet. If a task has a date, it should not drift.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Prep & Compliance"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    tasks = ensure_minimum_rows(
        model["prep_tasks"],
        {
            "due_date": None,
            "category": "",
            "task": "",
            "practical_note": "",
            "priority": "Medium",
            "status": "Not started",
            "owner": "",
            "source_url": "",
            "notes": "",
        },
    )
    end_row = 3 + len(tasks)
    add_list_validation(ws, f"E4:E{end_row}", ["High", "Medium", "Low"])
    add_list_validation(ws, f"F4:F{end_row}", ["Not started", "In progress", "Review", "Done"])
    for row_index, task in enumerate(tasks, start=4):
        ws.row_dimensions[row_index].height = 40
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", task["due_date"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"], number_format=DATE_FORMAT_SHORT if task["due_date"] else None)
        value_cell(ws, f"B{row_index}", task["category"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"C{row_index}", task["task"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", task["practical_note"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", task["priority"], fill=fill, font=FONTS["editable"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        status_fill, status_font = prep_status_style(task["status"])
        value_cell(ws, f"F{row_index}", task["status"], fill=status_fill, font=status_font, alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"G{row_index}", task["owner"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        url_font = FONTS["editable"] if task["source_url"].startswith("http") else FONTS["static_text"]
        value_cell(ws, f"H{row_index}", task["source_url"], fill=fill, font=url_font, alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"I{row_index}", task["notes"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def build_pack_sheet(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Pack List"], MERGE_SPECS["Pack List"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:H1", "Pack list", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:H2",
        "This is what should physically travel with you. Use 'Need to buy' if the item still has to be acquired first.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Pack List"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    items = ensure_minimum_rows(
        model["pack_items"],
        {"zone": "", "item": "", "who": "", "need_first": "", "priority": "Medium", "status": "To pack", "pack_note": "", "notes": ""},
    )
    end_row = 3 + len(items)
    add_list_validation(ws, f"E4:E{end_row}", ["High", "Medium", "Low"])
    add_list_validation(ws, f"F4:F{end_row}", ["To pack", "Need to buy", "Packed", "Clarify"])
    for row_index, item in enumerate(items, start=4):
        ws.row_dimensions[row_index].height = 34
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", item["zone"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"B{row_index}", item["item"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"C{row_index}", item["who"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", item["need_first"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", item["priority"], fill=fill, font=FONTS["editable"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        status_fill, status_font = pack_status_style(item["status"])
        value_cell(ws, f"F{row_index}", item["status"], fill=status_fill, font=status_font, alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"G{row_index}", item["pack_note"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"H{row_index}", item["notes"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def build_buy_sheet(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Buy List"], MERGE_SPECS["Buy List"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:J1", "Buy list", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:J2",
        "This is the clean shopping tracker. Split it into before-trip essentials and in-destination targets so nothing gets lost.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Buy List"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    items = ensure_minimum_rows(
        model["buy_items"],
        {
            "phase": "",
            "requester": "",
            "category": "",
            "item": "",
            "best_place_timing": "",
            "priority": "Medium",
            "status": "Need to buy",
            "budget": "",
            "why_it_matters": "",
            "notes": "",
        },
    )
    end_row = 3 + len(items)
    add_list_validation(ws, f"F4:F{end_row}", ["High", "Medium", "Low"])
    add_list_validation(
        ws,
        f"G4:G{end_row}",
        ["Need to buy", "Watch price", model["buy_in_destination_label"], "Bought", "Skip", "Clarify"],
    )
    for row_index, item in enumerate(items, start=4):
        ws.row_dimensions[row_index].height = 34
        fill = zebra_fill(row_index)
        for column_letter, key in zip(("A", "B", "C"), ("phase", "requester", "category")):
            value_cell(ws, f"{column_letter}{row_index}", item[key], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", item["item"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", item["best_place_timing"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"F{row_index}", item["priority"], fill=fill, font=FONTS["editable"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        status_fill, status_font = buy_status_style(item["status"], model["buy_in_destination_label"])
        value_cell(ws, f"G{row_index}", item["status"], fill=status_fill, font=status_font, alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"H{row_index}", item["budget"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["center"], border=BORDERS["row"])
        value_cell(ws, f"I{row_index}", item["why_it_matters"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"J{row_index}", item["notes"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def build_sources_sheet(ws, model: dict[str, Any]) -> None:
    apply_sheet_geometry(ws, WIDTH_SPECS["Sources"], MERGE_SPECS["Sources"])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    merge_value(ws, "A1:F1", "Sources used in this workbook", fill=FILLS["navy"], font=FONTS["title"], alignment=ALIGNMENTS["left_center"])
    merge_value(
        ws,
        "A2:F2",
        "Official URLs are kept in plain text so the facts can be re-checked later. User files and screenshots stay separate.",
        fill=FILLS["teal"],
        font=FONTS["subtitle"],
        alignment=ALIGNMENTS["left_center"],
    )
    for index, header in enumerate(HEADERS["Sources"], start=1):
        value_cell(
            ws,
            f"{get_column_letter(index)}3",
            header,
            fill=FILLS["teal"],
            font=FONTS["column_header"],
            alignment=ALIGNMENTS["center"],
            border=BORDERS["column_header"],
        )

    sources = ensure_minimum_rows(
        model["sources"],
        {
            "type": "",
            "source_name": "",
            "url_or_file_note": "",
            "used_for": "",
            "last_checked_or_file_date": "",
            "notes": "",
        },
    )
    for row_index, item in enumerate(sources, start=4):
        ws.row_dimensions[row_index].height = 32
        fill = zebra_fill(row_index)
        value_cell(ws, f"A{row_index}", item["type"], fill=fill, font=FONTS["static_label"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"B{row_index}", item["source_name"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        url_font = FONTS["editable"] if item["url_or_file_note"].startswith("http") else FONTS["static_text"]
        value_cell(ws, f"C{row_index}", item["url_or_file_note"], fill=fill, font=url_font, alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"D{row_index}", item["used_for"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"E{row_index}", item["last_checked_or_file_date"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])
        value_cell(ws, f"F{row_index}", item["notes"], fill=fill, font=FONTS["static_text"], alignment=ALIGNMENTS["left_top"], border=BORDERS["row"])


def dynamic_ranges(model: dict[str, Any]) -> dict[str, str]:
    bookings_end = max(4, 3 + max(1, len(model["bookings"])))
    prep_end = max(4, 3 + max(1, len(model["prep_tasks"])))
    pack_end = max(4, 3 + max(1, len(model["pack_items"])))
    buy_end = max(4, 3 + max(1, len(model["buy_items"])))
    return {
        "bookings_item_range": f"Bookings!B4:B{bookings_end}",
        "bookings_coverage_range": f"Bookings!K4:K{bookings_end}",
        "prep_status_range": f"'Prep & Compliance'!F4:F{prep_end}",
        "pack_status_range": f"'Pack List'!F4:F{pack_end}",
        "buy_status_range": f"'Buy List'!G4:G{buy_end}",
    }


def build_workbook(model: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    if getattr(wb, "calculation", None) is not None:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"

    for sheet_name in workbook_sheet_order(model):
        wb.create_sheet(title=sheet_name)

    ranges = dynamic_ranges(model)
    build_overview(wb["Overview"], model, ranges)
    build_bookings(wb["Bookings"], model, ranges)
    build_daily_plan(wb["Daily Plan"], model)
    build_options_sheet(wb[model["options_sheet_name"]], model)
    build_prep_sheet(wb["Prep & Compliance"], model)
    build_pack_sheet(wb["Pack List"], model)
    build_buy_sheet(wb["Buy List"], model)
    build_sources_sheet(wb["Sources"], model)
    return wb


def serialize_model(model: dict[str, Any]) -> dict[str, Any]:
    def convert(value: Any) -> Any:
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, list):
            return [convert(item) for item in value]
        if isinstance(value, dict):
            return {key: convert(item) for key, item in value.items()}
        return value

    return convert(model)


def main() -> None:
    args = parse_args()
    with Path(args.trip_model).expanduser().open("r", encoding="utf-8") as handle:
        raw_model = json.load(handle)
    model = normalize_trip_model(raw_model)
    workbook = build_workbook(model)
    output_path = choose_output_path(model, args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    if args.normalized_model_out:
        normalized_path = Path(args.normalized_model_out).expanduser()
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        with normalized_path.open("w", encoding="utf-8") as handle:
            json.dump(serialize_model(model), handle, indent=2)
            handle.write("\n")

    if not args.quiet:
        summary = {
            "output": str(output_path.resolve()),
            "sheet_order": workbook.sheetnames,
            "planner_days": (model["planner_end_date"] - model["planner_start_date"]).days + 1,
            "planned_adults": model["planned_adults"],
            "bookings": len(model["bookings"]),
            "daily_rows": len(model["daily_rows"]),
            "prep_tasks": len(model["prep_tasks"]),
            "pack_items": len(model["pack_items"]),
            "buy_items": len(model["buy_items"]),
            "sources": len(model["sources"]),
        }
        print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # pragma: no cover - CLI guard
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
